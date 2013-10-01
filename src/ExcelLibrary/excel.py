# -*- encoding: utf-8 -*-
# Copyright (c) 2013 Roman Merkushin
# rmerkushin@ya.ru

from xlrd import open_workbook
from openpyxl import load_workbook


class Excel(object):

    def __init__(self):
        self.wb = None
        self.file_format = None

    def open_workbook(self, file_name):
        """
        This function should be called before any other methods ExcelLibrary.
        `Open Workbook` provide opening Excel files in two formats: *.xls and *.xlsx.\n
        Example usage:
        | Open Workbook | C:\\test.xlsx |
        """
        if file_name.endswith(".xls"):
            self.wb = open_workbook(file_name, formatting_info=True)
            self.file_format = "xls"
        else:
            self.wb = load_workbook(file_name)
            self.file_format = "xlsx"

    def read_cell(self, sheet, row, column):
        """
        This function returns cell value.
        Sheets, rows and columns index start at 0.\n
        Example usage:
        | ${value} | Read Cell | 0 | 1 | 1 |
        `Read Cell` returns value from first sheet 'B2' cell.
        """
        if self.file_format == "xls":
            ws = self.wb.sheet_by_index(int(sheet))
            cell_value = ws.cell(int(row), int(column)).value
        else:
            ws = self.wb.worksheets[int(sheet)]
            cell_value = ws.cell(row=int(row), column=int(column)).value
        return cell_value
